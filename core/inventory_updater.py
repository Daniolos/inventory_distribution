"""Inventory update logic - generates updated Excel with post-distribution quantities."""

import io
from typing import BinaryIO

from openpyxl import load_workbook

from .models import TransferPreview, UpdatedInventoryResult


def apply_transfers_to_inventory(
    original_file: BinaryIO,
    previews: list[TransferPreview],
    source_column: str,
    header_row: int
) -> tuple[bytes, list[str]]:
    """
    Apply transfers to create updated inventory Excel.

    Args:
        original_file: Original uploaded Excel file
        previews: List of TransferPreview with planned transfers
        source_column: "Сток" or "Фото склад"
        header_row: 0-indexed header row

    Returns:
        Tuple of (Excel bytes, list of warnings)
    """
    warnings: list[str] = []

    # Load workbook preserving formatting
    original_file.seek(0)
    wb = load_workbook(original_file)
    ws = wb.active

    # Find column indices from header row (1-indexed in openpyxl)
    excel_header_row = header_row + 1  # Convert to 1-indexed
    column_map: dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=excel_header_row, column=col_idx).value
        if cell_value:
            column_map[str(cell_value).strip()] = col_idx

    # Find source column index
    source_col_idx = column_map.get(source_column)
    if not source_col_idx:
        warnings.append(f"Столбец источника '{source_column}' не найден")
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), warnings

    # Process each preview with transfers
    for preview in previews:
        if not preview.has_transfers:
            continue

        excel_row = preview.row_index

        # Calculate total quantity to subtract from source
        total_from_source = sum(t.quantity for t in preview.transfers)

        # Update source column (reduce)
        source_cell = ws.cell(row=excel_row, column=source_col_idx)
        original_value = source_cell.value if source_cell.value else 0
        try:
            original_value = int(float(original_value))
        except (ValueError, TypeError):
            original_value = 0

        new_value = original_value - total_from_source
        if new_value < 0:
            warnings.append(
                f"Строка {excel_row}: Исходное значение ({original_value}) < Перемещение ({total_from_source})"
            )
            new_value = 0
        source_cell.value = new_value

        # Update receiver columns (increase)
        for transfer in preview.transfers:
            receiver = transfer.receiver

            # Find receiver column
            receiver_col_idx = column_map.get(receiver)
            if not receiver_col_idx:
                # Try finding by store code prefix
                for col_name, col_idx in column_map.items():
                    if col_name.startswith(receiver.split()[0] if " " in receiver else receiver):
                        receiver_col_idx = col_idx
                        break

            if not receiver_col_idx:
                warnings.append(f"Строка {excel_row}: Столбец назначения '{receiver}' не найден")
                continue

            receiver_cell = ws.cell(row=excel_row, column=receiver_col_idx)
            receiver_value = receiver_cell.value if receiver_cell.value else 0
            try:
                receiver_value = int(float(receiver_value))
            except (ValueError, TypeError):
                receiver_value = 0

            receiver_cell.value = receiver_value + transfer.quantity

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), warnings


def generate_updated_inventory_result(
    original_file: BinaryIO,
    previews: list[TransferPreview],
    source_column: str,
    source_name: str,
    header_row: int
) -> UpdatedInventoryResult:
    """
    Generate UpdatedInventoryResult with updated inventory Excel.

    Args:
        original_file: Original uploaded Excel file
        previews: List of TransferPreview with planned transfers
        source_column: Column name ("Сток" or "Фото склад")
        source_name: Display name ("Сток" or "Фото")
        header_row: 0-indexed header row

    Returns:
        UpdatedInventoryResult with Excel bytes and metadata
    """
    from datetime import datetime

    data, warnings = apply_transfers_to_inventory(
        original_file,
        previews,
        source_column,
        header_row
    )

    # Count statistics
    rows_with_transfers = sum(1 for p in previews if p.has_transfers)
    total_quantity = sum(p.total_quantity for p in previews)

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"updated_inventory_{source_name}_{timestamp}.xlsx"

    return UpdatedInventoryResult(
        filename=filename,
        data=data,
        source_column=source_column,
        total_rows_updated=rows_with_transfers,
        total_quantity_transferred=total_quantity,
        warnings=warnings
    )
