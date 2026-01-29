#!/usr/bin/env python3
"""Generate test Excel file with defined scenarios for verification"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Store columns for convenience
STORE_COLS = [
    "125004 EKT-PC-Гринвич", "125005 EKT-PC-Мега", "125006 KZN-PC-Мега",
    "125007 MSK-PC-Гагаринский", "125008 MSK-PC-РИО Ленинский",
    "125009 NNV-PC-Фантастика", "125011 SPB-PC-Мега 2 Парнас",
    "125839 - MSK-PC-Outlet Белая Дача", "129877 MSK-PC-Мега 1 Теплый Стан",
    "130143 MSK-PCM-Мега 2 Химки", "150002 MSK-DV-Капитолий",
]

def all_stores_zero():
    """Helper: all stores = 0"""
    return {store: 0 for store in STORE_COLS}

def all_stores_value(val):
    """Helper: all stores = val"""
    return {store: val for store in STORE_COLS}

# Define test scenarios
scenarios = [
    # ==================== S1-S8: Original scenarios (with filter columns added) ====================
    {
        "name": "S1: Stock full, all stores empty",
        "description": "Stock=5, all 0 → distributes 1 to each (max 5 stores)",
        "Номенклатура": "Test Product A",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Shoes",
        **all_stores_zero(),
        "Сток": 5,
        "Фото склад": 0,
        "expected_script1": "125007=1, 125008=1, 129877=1, 130143=1, 150002=1 (priority order)",
        "expected_script2": "No change (no store >2)",
    },
    {
        "name": "S2: Stock full, some stores already have",
        "description": "Stock=3, 125007 and 125008 already have 1 → only distribute to empty",
        "Номенклатура": "Test Product B",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Bags",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 1,
        "125008 MSK-PC-РИО Ленинский": 1,
        "Сток": 3,
        "Фото склад": 0,
        "expected_script1": "129877=1, 130143=1, 150002=1 (skips 125007, 125008)",
        "expected_script2": "No change",
    },
    {
        "name": "S3: Stock empty",
        "description": "Stock=0 → nothing happens",
        "Номенклатура": "Test Product C",
        "Характеристика": "Size S",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Shoes",
        **all_stores_zero(),
        "Сток": 0,
        "Фото склад": 0,
        "expected_script1": "No distribution (Stock=0)",
        "expected_script2": "No change",
    },
    {
        "name": "S4: All stores full, Stock still has",
        "description": "Stock=3, all have 1 → Stock remains (nothing to distribute)",
        "Номенклатура": "Test Product D",
        "Характеристика": "Size XL",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Accessories",
        **all_stores_value(1),
        "Сток": 3,
        "Фото склад": 0,
        "expected_script1": "No distribution (all stores >0)",
        "expected_script2": "No change (none >2)",
    },
    {
        "name": "S5: One store has >2, others empty",
        "description": "125007=5, others=0 → distributes surplus (5-2=3) to empty",
        "Номенклатура": "Test Product E",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "",  # Edge case: empty collection
        "Наименование_доп": "Shoes",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 5,
        "Сток": 0,
        "Фото склад": 0,
        "expected_script1": "No change (Stock=0)",
        "expected_script2": "125007→125008=1, 125007→129877=1, 125007→130143=1",
    },
    {
        "name": "S6: Multiple stores >2, takes from highest first",
        "description": "125007=4, 125008=6 → takes from 125008 first (highest)",
        "Номенклатура": "Test Product F",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "",  # Edge case: empty additional name
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 4,
        "125008 MSK-PC-РИО Ленинский": 6,
        "Сток": 0,
        "Фото склад": 0,
        "expected_script1": "No change (Stock=0)",
        "expected_script2": "125008 distributes 4 (to 129877,130143,150002,125009), then 125007 distributes 2 (to 125011,125004)",
    },
    {
        "name": "S7: Surplus goes to Stock",
        "description": "125007=10, all others already have 1 → remainder to Stock",
        "Номенклатура": "Test Product G",
        "Характеристика": "Size S",
        "Коллекция (сезон)": "AW24",
        "Наименование_доп": "Bags",
        **all_stores_value(1),
        "125007 MSK-PC-Гагаринский": 10,
        "Сток": 0,
        "Фото склад": 0,
        "expected_script1": "No change (Stock=0, all >0)",
        "expected_script2": "125007→Сток=8 (surplus 10-2=8, no empty stores)",
    },
    {
        "name": "S8: Photo Stock Distribution",
        "description": "Photo Stock=3, Stock=0, all empty → distributes from Photo",
        "Номенклатура": "Test Product H",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "SS25",
        "Наименование_доп": "Accessories",
        **all_stores_zero(),
        "Сток": 0,
        "Фото склад": 3,
        "expected_script1": "With 'photo': 125007=1, 125008=1, 129877=1",
        "expected_script2": "No change",
    },

    # ==================== B1-B4: Minimum sizes rule (3 sizes) ====================
    # B1: Product with only 2 sizes in stock → should NOT transfer
    {
        "name": "B1a: MinSize 2 sizes - Size S",
        "description": "Only 2 sizes in stock → NO transfer (< 3 minimum)",
        "Номенклатура": "MinSize Product 2",  # Same product name!
        "Характеристика": "Size S",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "NO transfer - only 2 sizes available",
        "expected_script2": "No change",
    },
    {
        "name": "B1b: MinSize 2 sizes - Size M",
        "description": "Only 2 sizes in stock → NO transfer (< 3 minimum)",
        "Номенклатура": "MinSize Product 2",  # Same product name!
        "Характеристика": "Size M",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "NO transfer - only 2 sizes available",
        "expected_script2": "No change",
    },

    # B2: Product with exactly 3 sizes in stock → should transfer all 3
    {
        "name": "B2a: MinSize 3 sizes - Size S",
        "description": "Exactly 3 sizes → transfers 3 to first store",
        "Номенклатура": "MinSize Product 3",
        "Характеристика": "Size S",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Transfer to 125007 (3 sizes total)",
        "expected_script2": "No change",
    },
    {
        "name": "B2b: MinSize 3 sizes - Size M",
        "description": "Exactly 3 sizes → transfers 3 to first store",
        "Номенклатура": "MinSize Product 3",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Transfer to 125007 (3 sizes total)",
        "expected_script2": "No change",
    },
    {
        "name": "B2c: MinSize 3 sizes - Size L",
        "description": "Exactly 3 sizes → transfers 3 to first store",
        "Номенклатура": "MinSize Product 3",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Transfer to 125007 (3 sizes total)",
        "expected_script2": "No change",
    },

    # B3: Product with 5 sizes, store has 0 → only 3 should transfer
    {
        "name": "B3a: MinSize 5 sizes - Size XS",
        "description": "5 sizes available, store has 0 → only 3 transfer",
        "Номенклатура": "MinSize Product 5",
        "Характеристика": "Size XS",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Only 3 sizes to 125007 (not all 5)",
        "expected_script2": "No change",
    },
    {
        "name": "B3b: MinSize 5 sizes - Size S",
        "description": "5 sizes available, store has 0 → only 3 transfer",
        "Номенклатура": "MinSize Product 5",
        "Характеристика": "Size S",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Only 3 sizes to 125007 (not all 5)",
        "expected_script2": "No change",
    },
    {
        "name": "B3c: MinSize 5 sizes - Size M",
        "description": "5 sizes available, store has 0 → only 3 transfer",
        "Номенклатура": "MinSize Product 5",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Only 3 sizes to 125007 (not all 5)",
        "expected_script2": "No change",
    },
    {
        "name": "B3d: MinSize 5 sizes - Size L",
        "description": "5 sizes available, store has 0 → only 3 transfer",
        "Номенклатура": "MinSize Product 5",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Only 3 sizes to 125007 (not all 5)",
        "expected_script2": "No change",
    },
    {
        "name": "B3e: MinSize 5 sizes - Size XL",
        "description": "5 sizes available, store has 0 → only 3 transfer",
        "Номенклатура": "MinSize Product 5",
        "Характеристика": "Size XL",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Only 3 sizes to 125007 (not all 5)",
        "expected_script2": "No change",
    },

    # B4: Product where store already has 2+ sizes → normal rule applies
    {
        "name": "B4a: Normal rule - Size S (store has)",
        "description": "Store has 2+ sizes → normal rule (not 3-size rule)",
        "Номенклатура": "MinSize Product Normal",
        "Характеристика": "Size S",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 1,  # Store has this size
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Normal rule - L and XL transfer (store has 2+ sizes)",
        "expected_script2": "No change",
    },
    {
        "name": "B4b: Normal rule - Size M (store has)",
        "description": "Store has 2+ sizes → normal rule (not 3-size rule)",
        "Номенклатура": "MinSize Product Normal",
        "Характеристика": "Size M",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 1,  # Store has this size
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "Normal rule - L and XL transfer (store has 2+ sizes)",
        "expected_script2": "No change",
    },
    {
        "name": "B4c: Normal rule - Size L (store needs)",
        "description": "Store has 2+ sizes → normal rule, this size transfers",
        "Номенклатура": "MinSize Product Normal",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 0,  # Store needs this size
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "This size transfers to 125007",
        "expected_script2": "No change",
    },
    {
        "name": "B4d: Normal rule - Size XL (store needs)",
        "description": "Store has 2+ sizes → normal rule, this size transfers",
        "Номенклатура": "MinSize Product Normal",
        "Характеристика": "Size XL",
        "Коллекция (сезон)": "2024",
        "Наименование_доп": "Test MinSize",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 0,  # Store needs this size
        "Сток": 1,
        "Фото склад": 0,
        "expected_script1": "This size transfers to 125007",
        "expected_script2": "No change",
    },

    # ==================== D: Edge Cases ====================
    {
        "name": "D1: Decimal numbers",
        "description": "Stock=3.0, Store=1.0 → should be treated as integers",
        "Номенклатура": "Edge Case Decimal",
        "Характеристика": "Size M",
        "Коллекция (сезон)": 2024,  # Number instead of string
        "Наименование_доп": "Edge Cases",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 1.0,  # Decimal
        "Сток": 3.0,  # Decimal
        "Фото склад": 0,
        "expected_script1": "Normal processing (decimals → int)",
        "expected_script2": "No change",
    },
    {
        "name": "D2: Large numbers",
        "description": "Stock=999, Store=500 → normal processing",
        "Номенклатура": "Edge Case Large",
        "Характеристика": "Size L",
        "Коллекция (сезон)": "2025",
        "Наименование_доп": "Edge Cases",
        **all_stores_zero(),
        "125007 MSK-PC-Гагаринский": 500,
        "Сток": 999,
        "Фото склад": 0,
        "expected_script1": "Distributes to stores with 0",
        "expected_script2": "125007 surplus goes to other stores/Stock",
    },
]


def create_test_excel():
    wb = Workbook()

    # Sheet 1: Test Data (in format of real input file)
    ws_data = wb.active
    ws_data.title = "Test Input"

    # Header rows to match real file format
    ws_data['A1'] = ""
    ws_data['A2'] = "Параметры:"
    ws_data['C2'] = "Дата остатков: TEST"

    # Column headers (row 7 in real file, row 7 here)
    headers = [
        "Ид номенклатуры", "", "", "Ид характеристики",
        "Номенклатура", "Характеристика", "", "Коллекция (сезон)",
        "Наименование_доп", "Штрихкод", "",
        "125004 EKT-PC-Гринвич", "125005 EKT-PC-Мега", "125006 KZN-PC-Мега",
        "125007 MSK-PC-Гагаринский", "125008 MSK-PC-РИО Ленинский",
        "125009 NNV-PC-Фантастика", "125011 SPB-PC-Мега 2 Парнас",
        "125839 - MSK-PC-Outlet Белая Дача", "129877 MSK-PC-Мега 1 Теплый Стан",
        "130143 MSK-PCM-Мега 2 Химки", "150002 MSK-DV-Капитолий",
        "Сток", "Фото склад"
    ]

    for col, header in enumerate(headers, 1):
        ws_data.cell(row=7, column=col, value=header)
        ws_data.cell(row=7, column=col).font = Font(bold=True)

    # Sub-header row
    for col in range(12, 25):
        ws_data.cell(row=8, column=col, value="Остаток на складе")

    # Data rows
    store_cols = [
        "125004 EKT-PC-Гринвич", "125005 EKT-PC-Мега", "125006 KZN-PC-Мега",
        "125007 MSK-PC-Гагаринский", "125008 MSK-PC-РИО Ленинский",
        "125009 NNV-PC-Фантастика", "125011 SPB-PC-Мега 2 Парнас",
        "125839 - MSK-PC-Outlet Белая Дача", "129877 MSK-PC-Мега 1 Теплый Стан",
        "130143 MSK-PCM-Мега 2 Химки", "150002 MSK-DV-Капитолий",
        "Сток", "Фото склад"
    ]

    for i, scenario in enumerate(scenarios):
        row = 9 + i
        ws_data.cell(row=row, column=1, value=1000 + i)  # Ид номенклатуры
        ws_data.cell(row=row, column=4, value=2000 + i)  # Ид характеристики
        ws_data.cell(row=row, column=5, value=scenario["Номенклатура"])
        ws_data.cell(row=row, column=6, value=scenario["Характеристика"])

        # Filter columns
        ws_data.cell(row=row, column=8, value=scenario.get("Коллекция (сезон)", ""))
        ws_data.cell(row=row, column=9, value=scenario.get("Наименование_доп", ""))

        for j, store in enumerate(store_cols):
            val = scenario.get(store, 0)
            ws_data.cell(row=row, column=12 + j, value=val if val else None)

    # Sheet 2: Expected Results
    ws_expected = wb.create_sheet("Expected Results")
    ws_expected['A1'] = "Scenario"
    ws_expected['B1'] = "Description"
    ws_expected['C1'] = "Expected Script 1 (Stock→Stores)"
    ws_expected['D1'] = "Expected Script 2 (Balance)"

    for col in range(1, 5):
        ws_expected.cell(row=1, column=col).font = Font(bold=True)
        ws_expected.cell(row=1, column=col).fill = PatternFill("solid", fgColor="FFFF00")

    for i, scenario in enumerate(scenarios):
        row = 2 + i
        ws_expected.cell(row=row, column=1, value=scenario["name"])
        ws_expected.cell(row=row, column=2, value=scenario["description"])
        ws_expected.cell(row=row, column=3, value=scenario["expected_script1"])
        ws_expected.cell(row=row, column=4, value=scenario["expected_script2"])

    # Adjust column widths
    ws_expected.column_dimensions['A'].width = 40
    ws_expected.column_dimensions['B'].width = 50
    ws_expected.column_dimensions['C'].width = 50
    ws_expected.column_dimensions['D'].width = 60

    wb.save("test_scenarios.xlsx")
    print("Created: test_scenarios.xlsx")
    print(f"  - {len(scenarios)} test scenarios")
    print("  - Sheet 'Test Input': Input data matching real file format")
    print("  - Sheet 'Expected Results': Expected outcomes for verification")
    print()
    print("Scenario groups:")
    print("  - S1-S8: Original scenarios (with filter columns)")
    print("  - B1-B4: Minimum sizes rule (3 sizes)")
    print("  - D1-D2: Edge cases (decimals, large numbers)")

if __name__ == "__main__":
    create_test_excel()
