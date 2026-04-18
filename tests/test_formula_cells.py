"""Check behavior with formula cells in the source inventory.

If the real ERP export uses =SUM(...) for total-stock cells, openpyxl saves the
formula string. Our inventory_updater writes back ints — the question is whether
that interacts badly on re-read.
"""

import io
import pandas as pd
from openpyxl import Workbook

from core.distributor import StockDistributor
from core.models import DistributionConfig
from core.file_loader import find_header_row
from core.config import STOCK_COLUMN, PHOTO_STOCK_COLUMN, PRODUCT_NAME_COLUMN, VARIANT_COLUMN
from tests.conftest import STORE_COLS


def _make_workbook_with_total_row(
    product: str,
    variant_rows: list[tuple[str, int]],
    total_as_formula: bool = False,
    header_row: int = 1,
) -> bytes:
    """Workbook with a 'total' row at the top of a product group.

    If total_as_formula=True, the total cell contains =SUM(...) with the sizes' stocks.
    If False, it contains the raw summed number.
    """
    wb = Workbook()
    ws = wb.active
    columns = [PRODUCT_NAME_COLUMN, VARIANT_COLUMN, STOCK_COLUMN, PHOTO_STOCK_COLUMN] + STORE_COLS

    ws.cell(row=1, column=1, value="Отчёт по остаткам")
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=header_row + 1, column=col_idx, value=name)
    for col_idx, _ in enumerate(columns, start=1):
        ws.cell(row=header_row + 2, column=col_idx, value="Остаток на складе")

    # Total row (product-level aggregate)
    total_row = header_row + 3
    ws.cell(row=total_row, column=1, value=product)
    ws.cell(row=total_row, column=2, value="")  # empty variant
    total_value = sum(q for _, q in variant_rows)
    if total_as_formula:
        first_size_row = header_row + 4
        last_size_row = first_size_row + len(variant_rows) - 1
        ws.cell(row=total_row, column=3, value=f"=SUM(C{first_size_row}:C{last_size_row})")
    else:
        ws.cell(row=total_row, column=3, value=total_value)
    ws.cell(row=total_row, column=4, value=0)
    for i, store in enumerate(STORE_COLS):
        ws.cell(row=total_row, column=5 + i, value=0)

    # Variant rows
    for r_offset, (variant, stock) in enumerate(variant_rows):
        row = total_row + 1 + r_offset
        ws.cell(row=row, column=1, value=product)
        ws.cell(row=row, column=2, value=variant)
        ws.cell(row=row, column=3, value=stock)
        ws.cell(row=row, column=4, value=0)
        for i, store in enumerate(STORE_COLS):
            ws.cell(row=row, column=5 + i, value=0)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_raw_total_row_double_counts_stock():
    """If a raw 'total' row exists with the same product name, distributor treats
    it as an extra size — inflating the stock it can distribute.

    This test documents the behavior; it is a RED flag if real files have such rows.
    """
    wb_bytes = _make_workbook_with_total_row(
        product="Sweater",
        variant_rows=[("S", 10), ("M", 10), ("L", 10)],
        total_as_formula=False,
    )

    stream = io.BytesIO(wb_bytes)
    header_row, _ = find_header_row(stream)
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])

    # df has 4 rows: the total (stock=30) + 3 sizes (stock=10 each). Total visible = 60.
    total_visible_stock = int(df[STOCK_COLUMN].sum())
    assert total_visible_stock == 60, (
        f"Expected 60 visible stock (30 total + 3×10 sizes), got {total_visible_stock}"
    )

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )
    previews = StockDistributor(cfg).preview(df, "stock", header_row)
    total_transferred = sum(len(p.transfers) for p in previews)

    # If this prints a value > 30 (the real stock), we've confirmed the bug.
    print(f"\nReal product stock: 30, Visible (with phantom): 60, Transferred: {total_transferred}")
