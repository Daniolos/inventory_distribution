"""Hard invariant: total transfers across ALL passes must never exceed initial stock.

These tests stress the distributor with realistic catalogs (many products, many sizes,
lots of stores, multi-pass). Any failure here is a critical bug.
"""

import io
import pandas as pd
from openpyxl import Workbook

from core.distributor import StockDistributor
from core.models import DistributionConfig
from core.file_loader import find_header_row
from core.config import STOCK_COLUMN, PHOTO_STOCK_COLUMN, PRODUCT_NAME_COLUMN, VARIANT_COLUMN
from tests.conftest import STORE_COLS


def _make_workbook_bytes(rows, header_row=1):
    wb = Workbook()
    ws = wb.active
    columns = [PRODUCT_NAME_COLUMN, VARIANT_COLUMN, STOCK_COLUMN, PHOTO_STOCK_COLUMN] + STORE_COLS
    ws.cell(row=1, column=1, value="Отчёт по остаткам")
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=header_row + 1, column=col_idx, value=name)
    for col_idx, _ in enumerate(columns, start=1):
        ws.cell(row=header_row + 2, column=col_idx, value="Остаток на складе")
    for r_idx, row in enumerate(rows, start=header_row + 3):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(name, 0))
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _load(wb_bytes):
    stream = io.BytesIO(wb_bytes)
    header_row, _ = find_header_row(stream)
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])
    return df, header_row


def _pass(wb_bytes, cfg):
    df, header_row = _load(wb_bytes)
    distributor = StockDistributor(cfg)
    results = distributor.execute(df, "stock", header_row)
    inv = distributor.generate_updated_inventory(
        io.BytesIO(wb_bytes), df, "stock", header_row
    )
    items = sum(len(r.data) for r in results)
    return items, inv.total_quantity_transferred, inv.data


def _row(product, variant, stock):
    r = {PRODUCT_NAME_COLUMN: product, VARIANT_COLUMN: variant,
         STOCK_COLUMN: stock, PHOTO_STOCK_COLUMN: 0}
    for s in STORE_COLS:
        r[s] = 0
    return r


def test_realistic_catalog_stays_within_stock_single_pass():
    """Large catalog, many products, limited stock: total transfers ≤ initial stock."""
    rows = []
    for prod_idx in range(30):
        product = f"Sweater 244{prod_idx:03d}/1010 Black-42"
        for size in ["S", "M", "L", "XL", "XXL"]:
            rows.append(_row(product, size, stock=2))  # tight stock
    initial_stock = sum(r[STOCK_COLUMN] for r in rows)
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    items, qty, _ = _pass(wb_bytes, cfg)
    assert items == qty, f"execute items ({items}) must match inventory totals ({qty})"
    assert items <= initial_stock, (
        f"Transferred {items} > stock {initial_stock}. Duplicate transfers."
    )


def test_realistic_catalog_multi_pass_cumulative_within_stock():
    rows = []
    for prod_idx in range(30):
        product = f"Sweater 244{prod_idx:03d}/1010 Black-42"
        for size in ["S", "M", "L", "XL", "XXL"]:
            rows.append(_row(product, size, stock=2))
    initial_stock = sum(r[STOCK_COLUMN] for r in rows)
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    cumulative = 0
    for _ in range(5):
        items, _, wb_bytes = _pass(wb_bytes, cfg)
        cumulative += items

    assert cumulative <= initial_stock, (
        f"Cumulative transfers {cumulative} > stock {initial_stock}."
    )


def test_units_per_size_3_does_not_exceed_stock():
    """Even with units_per_size=3, distributor cannot exceed available stock per row."""
    rows = [_row("P", f"S{i}", stock=3) for i in range(5)]
    initial_stock = sum(r[STOCK_COLUMN] for r in rows)
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=3,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    items, qty, _ = _pass(wb_bytes, cfg)
    assert items == qty
    assert items <= initial_stock, f"{items} > {initial_stock}"


def test_phantom_total_row_with_empty_variant_inflates_stock_BUG():
    """Rows where variant is empty but stock is set act as phantom 'size' rows.

    If a catalog export includes product-level summary rows, the distributor
    treats them as real sizes and distributes their stock.
    """
    # Tight stock, enough stores that transfers from phantom visibly exceed real stock
    rows = [
        _row("Sweater", "", stock=100),    # phantom total row (looks like stock)
        _row("Sweater", "S", stock=1),
        _row("Sweater", "M", stock=1),
        _row("Sweater", "L", stock=1),
    ]
    real_stock = 3   # only the size rows are "real"
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    items, qty, _ = _pass(wb_bytes, cfg)
    assert items <= real_stock, (
        f"Transferred {items} items but REAL stock was only {real_stock}. "
        f"Phantom total row with empty variant is being distributed. "
        f"Fix: skip rows where variant is empty."
    )
