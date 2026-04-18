"""Integration test for multi-pass flow: execute, apply updates, re-run.

Verifies that running distributor a second time on the updated inventory
does NOT transfer the same stock again.
"""

import io
import pandas as pd
from openpyxl import Workbook

from core.distributor import StockDistributor
from core.models import DistributionConfig
from core.file_loader import find_header_row
from core.config import STOCK_COLUMN, PHOTO_STOCK_COLUMN, PRODUCT_NAME_COLUMN, VARIANT_COLUMN
from tests.conftest import STORE_COLS


def _make_workbook_bytes(rows: list[dict], header_row: int = 1) -> bytes:
    """Build an in-memory xlsx mirroring the real file layout.

    Layout (1-indexed Excel rows):
        1           preamble
        2           header (Номенклатура, Характеристика, Сток, Фото склад, stores...)
        3           sub-header (placeholder "Остаток на складе")
        4..         data rows
    """
    wb = Workbook()
    ws = wb.active

    columns = [PRODUCT_NAME_COLUMN, VARIANT_COLUMN, STOCK_COLUMN, PHOTO_STOCK_COLUMN] + STORE_COLS

    ws.cell(row=1, column=1, value="Отчёт по остаткам")
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=header_row + 1, column=col_idx, value=name)
    for col_idx, _ in enumerate(columns, start=1):
        ws.cell(row=header_row + 2, column=col_idx, value="Остаток на складе")

    for r_idx, row in enumerate(rows, start=header_row + 3):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(name, 0))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _load_df(workbook_bytes: bytes):
    stream = io.BytesIO(workbook_bytes)
    header_row, err = find_header_row(stream)
    assert err is None, err
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])
    return df, header_row


def _run_once(workbook_bytes: bytes, config: DistributionConfig):
    """Run preview + generate_updated_inventory once. Return (previews, new_bytes)."""
    df, header_row = _load_df(workbook_bytes)

    distributor = StockDistributor(config)
    previews = distributor.preview(df, source="stock", header_row=header_row)

    update_source = io.BytesIO(workbook_bytes)
    updated = distributor.generate_updated_inventory(
        update_source, df, source="stock", header_row=header_row
    )
    return previews, updated.data


def _make_row(product, variant, stock=0, store_quantities=None):
    row = {PRODUCT_NAME_COLUMN: product, VARIANT_COLUMN: variant,
           STOCK_COLUMN: stock, PHOTO_STOCK_COLUMN: 0}
    for s in STORE_COLS:
        row[s] = 0
    if store_quantities:
        row.update(store_quantities)
    return row


def test_second_run_does_not_duplicate_transfers():
    """After applying transfers, a second run on the updated inventory should
    find reduced stock and not re-transfer the same units."""
    rows = [
        _make_row("Sweater", "S", stock=10),
        _make_row("Sweater", "M", stock=10),
        _make_row("Sweater", "L", stock=10),
    ]
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        excluded_stores=[],
        balance_threshold=2,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    # Run 1
    previews1, updated_bytes = _run_once(wb_bytes, cfg)
    total1 = sum(len(p.transfers) for p in previews1)
    assert total1 > 0, "first run should produce transfers"

    # Run 2 on updated inventory
    previews2, _ = _run_once(updated_bytes, cfg)
    total2 = sum(len(p.transfers) for p in previews2)

    # Stock is 10 per size × 3 sizes = 30 units. With len(STORE_COLS)=7 stores and
    # units=1, each store gets 3 (one per size). Run 1 uses 7*3=21 units, leaves 9
    # (3 per size). Run 2 still has no store with ≥ target, so transfers
    # should be constrained by remaining stock — but NOT re-distribute the same units.
    assert total2 <= (30 - total1), (
        f"Second run transferred {total2} units but only {30 - total1} stock remained. "
        "This means transfers were duplicated."
    )


def test_total_transfers_never_exceed_stock():
    """Total transfers across any number of runs must not exceed initial stock."""
    rows = [
        _make_row("P", "S1", stock=5),
        _make_row("P", "S2", stock=5),
        _make_row("P", "S3", stock=5),
    ]
    initial_stock = sum(r[STOCK_COLUMN] for r in rows)
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        excluded_stores=[],
        balance_threshold=2,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    total_transferred = 0
    current = wb_bytes
    for _ in range(5):
        previews, current = _run_once(current, cfg)
        total_transferred += sum(len(p.transfers) for p in previews)

    assert total_transferred <= initial_stock, (
        f"Transferred {total_transferred} units but initial stock was only {initial_stock}. "
        "Bug: duplicate transfers across passes."
    )


def test_single_run_does_not_exceed_stock():
    """Sanity check: a single run cannot transfer more than available stock."""
    rows = [
        _make_row("P", "S1", stock=2),
        _make_row("P", "S2", stock=2),
        _make_row("P", "S3", stock=2),
    ]
    initial_stock = sum(r[STOCK_COLUMN] for r in rows)
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        excluded_stores=[],
        balance_threshold=2,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    previews, _ = _run_once(wb_bytes, cfg)
    total = sum(len(p.transfers) for p in previews)

    assert total <= initial_stock, f"Transferred {total} > stock {initial_stock}"


def test_updated_inventory_reflects_transfers():
    """After a run, the source column (Сток) in updated inventory should be reduced."""
    rows = [
        _make_row("P", "S1", stock=5),
        _make_row("P", "S2", stock=5),
        _make_row("P", "S3", stock=5),
    ]
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        excluded_stores=[],
        balance_threshold=2,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    previews, updated_bytes = _run_once(wb_bytes, cfg)
    total_from_source = sum(len(p.transfers) for p in previews)

    df_after, _ = _load_df(updated_bytes)
    total_stock_after = df_after[STOCK_COLUMN].sum()
    total_stock_before = 15

    assert total_stock_after == total_stock_before - total_from_source, (
        f"Stock should be {total_stock_before - total_from_source} after transferring "
        f"{total_from_source} units, but is {total_stock_after}"
    )
