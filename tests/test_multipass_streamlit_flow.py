"""Replicates the exact Streamlit flow: execute → generate_updated_inventory → persist bytes → next run.

This mirrors the button handler logic in app.py to catch any bug that only shows up
when both distributor.execute and distributor.generate_updated_inventory are called
back-to-back on the same df_filtered.
"""

import io
import pandas as pd
from openpyxl import Workbook

from core.distributor import StockDistributor
from core.models import DistributionConfig
from core.file_loader import find_header_row
from core.config import STOCK_COLUMN, PHOTO_STOCK_COLUMN, PRODUCT_NAME_COLUMN, VARIANT_COLUMN
from tests.conftest import STORE_COLS


def _make_workbook_bytes(rows, header_row=1):
    wb = Workbook()
    ws = wb.active
    columns = [PRODUCT_NAME_COLUMN, VARIANT_COLUMN, STOCK_COLUMN, PHOTO_STOCK_COLUMN] + STORE_COLS
    ws.cell(row=1, column=1, value="Отчёт по остаткам")
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=header_row + 1, column=col_idx, value=name)
    for col_idx, _ in enumerate(columns, start=1):
        ws.cell(row=header_row + 2, column=col_idx, value="Остаток на складе")
    for r_idx, row in enumerate(rows, start=header_row + 3):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(name, 0))
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _make_row(product, variant, stock=0):
    row = {PRODUCT_NAME_COLUMN: product, VARIANT_COLUMN: variant,
           STOCK_COLUMN: stock, PHOTO_STOCK_COLUMN: 0}
    for s in STORE_COLS:
        row[s] = 0
    return row


def _streamlit_like_pass(working_bytes, config):
    """Replicates the button handler in app.py.

    Loads df from working_bytes, calls distributor.execute AND distributor.generate_updated_inventory
    (both call preview internally), returns (new_results, new_inventory, new_working_bytes).
    """
    working_stream = io.BytesIO(working_bytes)
    header_row, err = find_header_row(working_stream)
    assert err is None
    working_stream.seek(0)
    df = pd.read_excel(working_stream, header=header_row, skiprows=[header_row + 1])

    df_filtered = df  # no UI filter in tests

    distributor = StockDistributor(config)
    new_results = distributor.execute(df_filtered, "stock", header_row)

    update_source = io.BytesIO(working_bytes)
    new_inventory = distributor.generate_updated_inventory(
        update_source, df_filtered, "stock", header_row
    )
    return new_results, new_inventory, new_inventory.data


def test_two_passes_stay_within_stock_budget():
    """Two successive passes via the exact Streamlit flow must not duplicate."""
    initial_stock_per_row = 10
    rows = [
        _make_row("Sweater", "S", stock=initial_stock_per_row),
        _make_row("Sweater", "M", stock=initial_stock_per_row),
        _make_row("Sweater", "L", stock=initial_stock_per_row),
    ]
    total_stock = 3 * initial_stock_per_row
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    # Pass 1
    r1, inv1, wb_bytes = _streamlit_like_pass(wb_bytes, cfg)
    items1 = sum(len(r.data) for r in r1)
    qty1 = inv1.total_quantity_transferred
    assert items1 == qty1, f"execute items ({items1}) != inventory totals ({qty1})"

    # Pass 2 on updated inventory
    r2, inv2, wb_bytes = _streamlit_like_pass(wb_bytes, cfg)
    items2 = sum(len(r.data) for r in r2)
    qty2 = inv2.total_quantity_transferred

    # Total transferred across both passes must not exceed initial stock
    assert items1 + items2 <= total_stock, (
        f"Pass 1: {items1}, Pass 2: {items2}, Total: {items1 + items2} > stock {total_stock}. "
        "Duplicate transfers across passes."
    )


def test_identical_df_filtered_passed_twice_gives_consistent_results():
    """Calling execute and generate_updated_inventory on the same df must agree."""
    rows = [_make_row("P", f"S{i}", stock=3) for i in range(5)]
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    stream = io.BytesIO(wb_bytes)
    header_row, _ = find_header_row(stream)
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])

    distributor = StockDistributor(cfg)
    results = distributor.execute(df, "stock", header_row)
    inv = distributor.generate_updated_inventory(
        io.BytesIO(wb_bytes), df, "stock", header_row
    )

    items_from_execute = sum(len(r.data) for r in results)
    assert items_from_execute == inv.total_quantity_transferred


def test_df_filtered_is_not_mutated_by_execute():
    """After distributor.execute, df_filtered values must be unchanged."""
    rows = [_make_row("P", f"S{i}", stock=3) for i in range(5)]
    wb_bytes = _make_workbook_bytes(rows)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    stream = io.BytesIO(wb_bytes)
    header_row, _ = find_header_row(stream)
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])

    before = df[STOCK_COLUMN].tolist()

    distributor = StockDistributor(cfg)
    distributor.execute(df, "stock", header_row)

    after = df[STOCK_COLUMN].tolist()
    assert before == after, "df stock values changed after execute()"


def test_generate_updated_inventory_does_not_mutate_source_bytes():
    """Input bytes handed into generate_updated_inventory must be unchanged."""
    rows = [_make_row("P", f"S{i}", stock=3) for i in range(5)]
    wb_bytes = _make_workbook_bytes(rows)
    wb_bytes_copy = bytes(wb_bytes)

    cfg = DistributionConfig(
        store_priority=STORE_COLS,
        target_sizes_filled=3,
        units_per_size=1,
        min_product_sizes=1,
        max_product_sizes=99,
    )

    stream = io.BytesIO(wb_bytes)
    header_row, _ = find_header_row(stream)
    stream.seek(0)
    df = pd.read_excel(stream, header=header_row, skiprows=[header_row + 1])

    distributor = StockDistributor(cfg)
    distributor.generate_updated_inventory(
        io.BytesIO(wb_bytes), df, "stock", header_row
    )

    assert wb_bytes == wb_bytes_copy, "Source bytes mutated by generate_updated_inventory"
